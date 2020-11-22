import pandas as pd
from openpyxl import load_workbook


def append_df_to_excel(filename, df, deduplicate_column=None, skiprows=0, **to_excel_kwargs):
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = load_workbook(filename)
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    sheet_name = writer.book.sheetnames[0]

    df_prev = pd.read_excel(filename, skiprows=skiprows)[df.columns].dropna(how='all')
    startrow = len(df_prev) + skiprows + 1

    if deduplicate_column:
        df = df[~df[deduplicate_column].isin(df_prev.to_numpy().flatten())]

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()
